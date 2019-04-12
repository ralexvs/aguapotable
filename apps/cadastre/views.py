from django.shortcuts import render
from django.views.generic.base import TemplateView
from django.views.generic.list import ListView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic.detail import DetailView
from django.urls import reverse, reverse_lazy
from django.contrib.admin.views.decorators import staff_member_required
from django.utils.decorators import method_decorator
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.shortcuts import redirect
from .models import (
    Subscribers, 
    Measurer,
    Cadastral,
    )

from .forms import (
    SubscribersForm, 
    MeasurerForm, 
    CadastralForm,
    )


# Create your views here.

class StaffRequieredMixin(object):
    """
    Este mixin requerirá que el usuario sea miembro del staff
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_staff:
            return redirect(reverse_lazy('login'))
    
        return super(StaffRequieredMixin, self).dispatch(request, *args, **kwargs)
    

#ListView
class SubscribersListView(ListView):
    model = Subscribers
    #template_name = "cadastre/list_subscribers.html"
    #paginate_by = 10
    
class MeasurerListView(ListView):
    model = Measurer
    #paginate_by = 10

class CadastralListView(ListView):
    model = Cadastral


#DetailView

class SubscribersDetailView(DetailView):
    model = Subscribers

class MeasurerDetailView(DetailView):
    model = Measurer

class CadastralDetailView(DetailView):
    model = Cadastral


#CreateView
@method_decorator(staff_member_required, name='dispatch')
class SubscribersCreateView(CreateView):
    model = Subscribers
    form_class = SubscribersForm
    template_name = "cadastre/subscribers_form.html"
    success_url = reverse_lazy('catastro:subscribers_list')

    
    
@method_decorator(staff_member_required, name='dispatch')
class MeasurerCreateView(CreateView):
    model = Measurer
    form_class = MeasurerForm
    template_name = "cadastre/measurer_form.html"
    success_url = reverse_lazy('catastro:measurer_list')

@method_decorator(staff_member_required, name='dispatch')
class CadastralCreateView(CreateView):
    model = Cadastral
    form_class = CadastralForm
    template_name = "cadastre/cadastral_form.html"
    success_url = reverse_lazy('catastro:cadastral_create')



#UpdateView

@method_decorator(staff_member_required, name='dispatch')    
class SubscribersUpdateView(UpdateView):
    model = Subscribers
    form_class = SubscribersForm
    template_name_suffix  = "_update_form"
    
    def get_success_url(self):
        return reverse_lazy('catastro:subscribers_update', args=[self.object.id]) + '?ok'

@method_decorator(staff_member_required, name='dispatch')
class MeasurerUpdateView(UpdateView):
    model = Measurer
    form_class = MeasurerForm
    template_name_suffix = "_update_form"

    def get_success_url(self):
        return reverse_lazy('catastro:measurer_update', args=[self.object.id]) + '?ok'

@method_decorator(staff_member_required, name='dispatch')
class CadastralUpdateView(UpdateView):
    model = Cadastral
    form_class = CadastralForm
    template_name_suffix = "_update_form"
    
    def get_success_url(self):
        return reverse_lazy('catastro:cadastral_update', args=[self.object.id]) + '?ok'




#DeleteView

@method_decorator(staff_member_required, name='dispatch')
class SubscribersDeleteView(DeleteView):
    model = Subscribers
    success_url = reverse_lazy('catastro:subscribers_list')

@method_decorator(staff_member_required, name='dispatch')
class MeasurerDeleteView(DeleteView):
    model = Measurer
    success_url = reverse_lazy('catastro:measurer_list')

@method_decorator(staff_member_required, name='dispatch')
class CadastralDeleteView(DeleteView):
    model = Cadastral
    success_url = reverse_lazy('catastro:cadastral_list')

#Reportes
class ReportSubscribersExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        subscribers = Subscribers.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1']= 'REPORTE DE ABONADOS'
        ws.merge_cells('B1:K1')
        ws['B3'] ='Item'
        ws['C3'] ='Identificación'
        ws['D3'] ='Apellidos'
        ws['E3'] ='Nombres'
        ws['F3'] ='Dirección'
        ws['G3'] ='Email'
        ws['H3'] ='Teléfono'
        ws['I3'] ='Celular'
        ws['J3'] ='Estado'
        ws['K3'] ='Nació'
        ws['B1'].font = Font(bold=True, size=16)
        ws['B3'].font = Font(bold=True)
        ws['C3'].font = Font(bold=True)
        ws['D3'].font = Font(bold=True)
        ws['E3'].font = Font(bold=True)
        ws['F3'].font = Font(bold=True)
        ws['G3'].font = Font(bold=True)
        ws['H3'].font = Font(bold=True)
        ws['I3'].font = Font(bold=True)
        ws['J3'].font = Font(bold=True)
        ws['K3'].font = Font(bold=True)
        ws['B1'].alignment = Alignment(horizontal="center")
        ws['C3'].alignment = Alignment(horizontal="center")
        ws['D3'].alignment = Alignment(horizontal="center")
        ws['E3'].alignment = Alignment(horizontal="center")
        ws['F3'].alignment = Alignment(horizontal="center")
        ws['G3'].alignment = Alignment(horizontal="center")
        ws['H3'].alignment = Alignment(horizontal="center")
        ws['I3'].alignment = Alignment(horizontal="center")
        ws['J3'].alignment = Alignment(horizontal="center")
        ws['K3'].alignment = Alignment(horizontal="center")
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
        ws.column_dimensions['K'].width = 15

        counter = 4
        for subscriber in subscribers:
            ws.cell(row = counter, column = 2).value = subscriber.id
            ws.cell(row = counter, column = 3).value = subscriber.identification
            ws.cell(row = counter, column = 4).value = subscriber.surname
            ws.cell(row = counter, column = 5).value = subscriber.name
            ws.cell(row = counter, column = 6).value = subscriber.address
            ws.cell(row = counter, column = 7).value = subscriber.email
            ws.cell(row = counter, column = 8).value = subscriber.phone
            ws.cell(row = counter, column = 9).value = subscriber.mobile
            ws.cell(row = counter, column = 10).value = subscriber.state
            ws.cell(row = counter, column = 11).value = subscriber.birth_date
            counter+=1
        #define el nombre delñ archivo
        file_name = "RporteAbonadosExcel.xlsx"
        #Definir el tipo de respuesta que va a dar
        response = HttpResponse(content_type = "application/ms_excel")
        content = "attachment; filename = {0}".format(file_name)
        response['Content-Disposition'] = content
        wb.save(response)
        return response 
